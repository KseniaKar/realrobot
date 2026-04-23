from __future__ import annotations

import csv
import re
from pathlib import Path

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent
MOSCOW_DIR = BASE_DIR / "moscow"
XLSX_PATH = MOSCOW_DIR / "2026-04.xlsx"
NORMALIZED_PATH = MOSCOW_DIR / "2026-04.normalized.csv"
WITH_NORM_PATH = MOSCOW_DIR / "2026-04.with_norm.csv"
STATS_PATH = BASE_DIR / "matches" / "moscow_2026_04_new_organizations_stats.csv"
NEW_ORGS_PATH = BASE_DIR / "matches" / "moscow_2026_04_new_organizations.csv"

OUTPUT_COLUMNS = [
    "id",
    "name",
    "region",
    "city",
    "district",
    "address",
    "phone",
    "email",
    "site",
    "rubric",
    "subrubric",
    "lon",
    "lat",
]

CANONICAL_COLUMNS = {
    "id": ["ID"],
    "name": ["Название"],
    "region": ["Регион"],
    "city": ["Город"],
    "district": ["Район города", "Район"],
    "address": ["Адрес"],
    "phone": ["Телефон"],
    "email": ["Email"],
    "site": ["Сайт"],
    "rubric": ["Рубрика"],
    "subrubric": ["Подрубрика"],
    "lon": ["Долгота"],
    "lat": ["Широта"],
}

WS_RE = re.compile(r"\s+")
PUNCT_RE = re.compile(r"[\"'`]")
NON_WORD_RE = re.compile(r"[^0-9a-zа-яё/]+", re.IGNORECASE)
TRIM_TAIL_RE = re.compile(
    r"(?:,?\s*(?:этаж|помещение|комната|офис|кабинет|антресоль|подвал|цоколь|чердак|мансарда|квартира)\b.*)$",
    re.IGNORECASE,
)

REPLACEMENTS = [
    (r"\bроссийская федерация\b", " "),
    (r"\bвнутригородская территория муниципальный округ\b", " "),
    (r"\bмуниципальный округ\b", " "),
    (r"\bгород федерального значения\b", " "),
    (r"\bг(?:ород)?\.?\s*москва\b", " "),
    (r"\bг(?:ород)?\b", " "),
    (r"\bмосква\b", " "),
    (r"\b(?:центральный|северный|северо-восточный|восточный|юго-восточный|южный|юго-западный|западный|северо-западный|зеленоградский|троицкий|новомосковский)\s+административный\s+округ\b", " "),
    (r"\bрайон\b", " "),
    (r"\bпоселение\b", " "),
    (r"\bтерритория\b", " "),
    (r"\bул(?:ица)?\b", "улица"),
    (r"\bпр-?кт\b", "проспект"),
    (r"\bпр-т\b", "проспект"),
    (r"\bпер\b", "переулок"),
    (r"\bбул\b", "бульвар"),
    (r"\bнаб\.?\b", "набережная"),
    (r"\bпл\b", "площадь"),
    (r"\bш\b", "шоссе"),
    (r"\bтуп\b", "тупик"),
    (r"\bпр\b", "проезд"),
    (r"\bд(?:ом)?\b", "дом"),
    (r"\bк(?:орпус)?\b", "корпус"),
    (r"\bстр(?:оение)?\b", "строение"),
    (r"\bсоор(?:ужение)?\b", "сооружение"),
    (r"\bвл\b", "владение"),
]


def normalize_address(value: object) -> str:
    text = ("" if value is None else str(value)).strip().lower().replace("ё", "е")
    text = text.replace("№", " ")
    text = PUNCT_RE.sub(" ", text)
    text = text.replace("\\", "/")
    text = TRIM_TAIL_RE.sub("", text)
    for pattern, replacement in REPLACEMENTS:
        text = re.sub(pattern, replacement, text, flags=re.IGNORECASE)
    text = NON_WORD_RE.sub(" ", text)
    text = re.sub(r"\bстроение\s+([0-9]+)\b", r"стр \1", text)
    text = re.sub(r"\bкорпус\s+([0-9a-zа-я]+)\b", r"корп \1", text)
    text = re.sub(r"\bдом\s+([0-9a-zа-я/]+)\b", r"\1", text)
    text = re.sub(r"\bвладение\s+([0-9a-zа-я/]+)\b", r"вл \1", text)
    return WS_RE.sub(" ", text).strip(" ,")


def first_present(row: dict[str, object], candidates: list[str]) -> object:
    for key in candidates:
        value = row.get(key)
        if value not in (None, ""):
            return value
    return ""


def normalize_row(row: dict[str, object]) -> dict[str, str]:
    out: dict[str, str] = {}
    for target, candidates in CANONICAL_COLUMNS.items():
        value = first_present(row, candidates)
        out[target] = "" if value is None else str(value)
    return out


def stream_xlsx_rows(path: Path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = worksheet.iter_rows(values_only=True)
    header = [str(cell).strip() if cell is not None else "" for cell in next(rows)]
    for values in rows:
        yield {header[idx]: values[idx] for idx in range(min(len(header), len(values)))}
    workbook.close()


def write_snapshot_csvs() -> int:
    count = 0
    WITH_NORM_PATH.parent.mkdir(parents=True, exist_ok=True)
    with NORMALIZED_PATH.open("w", encoding="utf-8-sig", newline="") as normalized_f, WITH_NORM_PATH.open(
        "w", encoding="utf-8-sig", newline=""
    ) as with_norm_f:
        normalized_writer = csv.DictWriter(normalized_f, fieldnames=OUTPUT_COLUMNS)
        with_norm_writer = csv.DictWriter(with_norm_f, fieldnames=OUTPUT_COLUMNS + ["address_norm"])
        normalized_writer.writeheader()
        with_norm_writer.writeheader()
        for raw_row in stream_xlsx_rows(XLSX_PATH):
            row = normalize_row(raw_row)
            if not row["address"]:
                continue
            normalized_writer.writerow(row)
            row["address_norm"] = normalize_address(row["address"])
            with_norm_writer.writerow(row)
            count += 1
    return count


def load_ids(path: Path) -> set[str]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return {(row.get("id") or "").strip() for row in csv.DictReader(f) if (row.get("id") or "").strip()}


def write_new_org_stats() -> tuple[int, int, int]:
    previous_ids = load_ids(MOSCOW_DIR / "2026-03.with_norm.csv")
    total_rows = 0
    new_rows: list[dict[str, str]] = []
    with WITH_NORM_PATH.open("r", encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            total_rows += 1
            row_id = (row.get("id") or "").strip()
            if row_id and row_id not in previous_ids:
                new_rows.append(row)

    NEW_ORGS_PATH.parent.mkdir(parents=True, exist_ok=True)
    with NEW_ORGS_PATH.open("w", encoding="utf-8-sig", newline="") as f:
        fieldnames = OUTPUT_COLUMNS + ["address_norm"]
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(new_rows)

    stats: dict[tuple[str, str], int] = {}
    for row in new_rows:
        district = (row.get("district") or "Не указано").strip() or "Не указано"
        rubric = (row.get("rubric") or "Не указано").strip() or "Не указано"
        stats[(district, rubric)] = stats.get((district, rubric), 0) + 1

    with STATS_PATH.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["district", "rubric", "new_organizations"])
        writer.writeheader()
        for (district, rubric), count in sorted(stats.items(), key=lambda item: item[1], reverse=True):
            writer.writerow({"district": district, "rubric": rubric, "new_organizations": count})

    return total_rows, len(previous_ids), len(new_rows)


def main() -> None:
    if NORMALIZED_PATH.exists() and WITH_NORM_PATH.exists():
        converted_rows = sum(1 for _ in WITH_NORM_PATH.open("r", encoding="utf-8-sig")) - 1
    else:
        converted_rows = write_snapshot_csvs()
    total_rows, previous_ids_count, new_rows = write_new_org_stats()
    print(
        {
            "xlsx": str(XLSX_PATH),
            "converted_rows": converted_rows,
            "total_2026_04_rows": total_rows,
            "previous_2026_03_ids": previous_ids_count,
            "new_vs_2026_03_ids": new_rows,
            "normalized_csv": str(NORMALIZED_PATH),
            "with_norm_csv": str(WITH_NORM_PATH),
            "new_orgs_csv": str(NEW_ORGS_PATH),
            "stats_csv": str(STATS_PATH),
        }
    )


if __name__ == "__main__":
    main()
