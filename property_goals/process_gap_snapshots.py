"""
Process the new gap-filling 2GIS snapshots (2022–2024) into .with_norm.csv format.

New snapshots:
  Flat xlsx : 2022-01, 2022-05, 2022-08, 2022-12, 2023-12
  District folders: 202405 -> 2024-05, 202409 -> 2024-09, 202410 -> 2024-10

Output: moscow/{label}.with_norm.csv for each new label.
Skips labels whose .with_norm.csv already exists.
"""

from __future__ import annotations

import csv
import re
from pathlib import Path

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent
MOSCOW_DIR = BASE_DIR / "moscow"

FLAT_LABELS = ["2022-01", "2022-05", "2022-08", "2022-12", "2023-12"]

# folder_name -> output label
DIR_LABELS: dict[str, str] = {
    "202405": "2024-05",
    "202409": "2024-09",
    "202410": "2024-10",
    "202505": "2025-05",
    "202507": "2025-07",
}

OUTPUT_COLUMNS = ["id", "name", "address", "address_norm", "rubric", "subrubric", "lat", "lon"]

CANONICAL_COLUMNS: dict[str, list[str]] = {
    "id":       ["ID"],
    "name":     ["Название"],
    "address":  ["Адрес"],
    "rubric":   ["Рубрика"],
    "subrubric":["Подрубрика"],
    "lat":      ["Широта"],
    "lon":      ["Долгота"],
}

# ── Address normalisation (identical to process_2026_04_snapshot.py) ─────────
WS_RE        = re.compile(r"\s+")
PUNCT_RE     = re.compile(r"[\"'`]")
NON_WORD_RE  = re.compile(r"[^0-9a-zа-яё/]+", re.IGNORECASE)
TRIM_TAIL_RE = re.compile(
    r"(?:,?\s*(?:этаж|помещение|комната|офис|кабинет|антресоль|подвал|цоколь"
    r"|чердак|мансарда|квартира)\b.*)$",
    re.IGNORECASE,
)
REPLACEMENTS = [
    (r"\bроссийская федерация\b",                                             " "),
    (r"\bвнутригородская территория муниципальный округ\b",                   " "),
    (r"\bмуниципальный округ\b",                                              " "),
    (r"\bгород федерального значения\b",                                      " "),
    (r"\bг(?:ород)?\.?\s*москва\b",                                           " "),
    (r"\bг(?:ород)?\b",                                                       " "),
    (r"\bмосква\b",                                                           " "),
    (r"\b(?:центральный|северный|северо-восточный|восточный|юго-восточный"
     r"|южный|юго-западный|западный|северо-западный|зеленоградский"
     r"|троицкий|новомосковский)\s+административный\s+округ\b",              " "),
    (r"\bрайон\b",    " "), (r"\bпоселение\b", " "), (r"\bтерритория\b", " "),
    (r"\bул(?:ица)?\b", "улица"),   (r"\bпр-?кт\b",    "проспект"),
    (r"\bпр-т\b",       "проспект"),(r"\bпер\b",        "переулок"),
    (r"\bбул\b",        "бульвар"), (r"\bнаб\.?\b",     "набережная"),
    (r"\bпл\b",         "площадь"), (r"\bш\b",          "шоссе"),
    (r"\bтуп\b",        "тупик"),   (r"\bпр\b",         "проезд"),
    (r"\bд(?:ом)?\b",   "дом"),     (r"\bк(?:орпус)?\b","корпус"),
    (r"\bстр(?:оение)?\b","строение"),(r"\bсоор(?:ужение)?\b","сооружение"),
    (r"\bвл\b",         "владение"),
]


def normalize_address(value: object) -> str:
    text = ("" if value is None else str(value)).strip().lower().replace("ё", "е")
    text = text.replace("№", " ")
    text = PUNCT_RE.sub(" ", text)
    text = text.replace("\\", "/")
    text = TRIM_TAIL_RE.sub("", text)
    for pattern, replacement in REPLACEMENTS:
        text = re.sub(pattern, replacement, text, flags=re.IGNORECASE)
    text = NON_WORD_RE.sub(" ", text)
    text = re.sub(r"\bстроение\s+([0-9]+)\b",         r"стр \1",  text)
    text = re.sub(r"\bкорпус\s+([0-9a-zа-я]+)\b",     r"корп \1", text)
    text = re.sub(r"\bдом\s+([0-9a-zа-я/]+)\b",       r"\1",      text)
    text = re.sub(r"\bвладение\s+([0-9a-zа-я/]+)\b",  r"вл \1",   text)
    return WS_RE.sub(" ", text).strip(" ,")


def first_present(row: dict, candidates: list[str]) -> str:
    for key in candidates:
        value = row.get(key)
        if value not in (None, ""):
            return str(value)
    return ""


def normalize_row(row: dict) -> dict[str, str]:
    return {target: first_present(row, cands) for target, cands in CANONICAL_COLUMNS.items()}


def stream_xlsx_rows(path: Path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = worksheet.iter_rows(values_only=True)
    header = [str(cell).strip() if cell is not None else "" for cell in next(rows)]
    for values in rows:
        yield {header[i]: values[i] for i in range(min(len(header), len(values)))}
    workbook.close()


def find_district_subfolder(folder: Path) -> Path:
    """Return the single subdirectory that contains the actual xlsx files."""
    subdirs = [d for d in folder.iterdir() if d.is_dir()]
    return subdirs[0] if len(subdirs) == 1 else folder


def process_snapshot(label: str, sources: list[Path]) -> None:
    out_path = MOSCOW_DIR / f"{label}.with_norm.csv"
    if out_path.exists():
        row_count = sum(1 for _ in out_path.open("r", encoding="utf-8-sig")) - 1
        print(f"  {label}: already exists ({row_count} rows) — skipping")
        return

    count = 0
    with out_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=OUTPUT_COLUMNS)
        writer.writeheader()
        for src in sources:
            for raw_row in stream_xlsx_rows(src):
                row = normalize_row(raw_row)
                if not row["address"]:
                    continue
                row["address_norm"] = normalize_address(row["address"])
                writer.writerow(row)
                count += 1

    print(f"  {label}: {count} rows -> {out_path.name}")


def main() -> None:
    print("=== Processing gap snapshots ===\n")

    print("Flat xlsx files:")
    for label in FLAT_LABELS:
        xlsx_path = MOSCOW_DIR / f"{label}.xlsx"
        if not xlsx_path.exists():
            print(f"  {label}: xlsx not found — skipping")
            continue
        process_snapshot(label, [xlsx_path])

    print("\nDistrict-folder snapshots:")
    for folder_name, label in DIR_LABELS.items():
        folder = MOSCOW_DIR / folder_name
        if not folder.exists():
            print(f"  {label}: folder {folder_name}/ not found — skipping")
            continue
        subfolder = find_district_subfolder(folder)
        sources = sorted(subfolder.glob("*.xlsx"))
        print(f"  {label}: merging {len(sources)} files from {folder_name}/{subfolder.name}/")
        process_snapshot(label, sources)

    print("\nDone.")


if __name__ == "__main__":
    main()
