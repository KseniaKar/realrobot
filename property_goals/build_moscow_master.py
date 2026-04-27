from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent
MOSCOW_DIR = BASE_DIR / "moscow"
OUTPUT_PATH = BASE_DIR / "moscow_master.csv"

OUTPUT_COLUMNS = [
    "source_file",
    "source_date",
    "source_format",
    "id",
    "name",
    "region",
    "city",
    "district",
    "address",
    "postal_code",
    "phone",
    "mobile_phone",
    "email",
    "site",
    "rubric",
    "subrubric",
    "rating",
    "reviews_count",
    "ratings_count",
    "working_hours",
    "payment_methods",
    "lat",
    "lon",
]

CANONICAL_COLUMNS = {
    "id": ["ID"],
    "name": ["Название"],
    "region": ["Регион"],
    "city": ["Город"],
    "district": ["Район", "Район города"],
    "address": ["Адрес"],
    "postal_code": ["Индекс"],
    "phone": ["Телефон"],
    "mobile_phone": ["Мобильный телефон"],
    "email": ["Email", "Email с сайта"],
    "site": ["Сайт"],
    "rubric": ["Рубрика"],
    "subrubric": ["Подрубрика"],
    "rating": ["Рейтинг"],
    "reviews_count": ["Кол-во отзывов"],
    "ratings_count": ["Кол-во оценок"],
    "working_hours": ["Время работы"],
    "payment_methods": ["Способы оплаты"],
    "lat": ["Широта"],
    "lon": ["Долгота"],
}


def parse_source_date(path: Path) -> str:
    return f"{path.stem}-01"


def first_present(row: dict[str, object], candidates: list[str]) -> object:
    for key in candidates:
        value = row.get(key)
        if value not in (None, ""):
            return value
    return None


def normalize_row(row: dict[str, object], path: Path) -> dict[str, object]:
    out: dict[str, object] = {
        "source_file": path.name,
        "source_date": parse_source_date(path),
        "source_format": path.suffix.lower().lstrip("."),
    }

    for target, candidates in CANONICAL_COLUMNS.items():
        out[target] = first_present(row, candidates)

    return out


def stream_xlsx_rows(path: Path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = worksheet.iter_rows(values_only=True)
    header = [str(cell).strip() if cell is not None else "" for cell in next(rows)]

    for values in rows:
        row = {header[idx]: values[idx] for idx in range(min(len(header), len(values)))}
        yield row

    workbook.close()


def stream_csv_rows(path: Path):
    with open(path, "r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle, delimiter=";")
        for row in reader:
            yield row


def main() -> None:
    files = sorted([path for path in MOSCOW_DIR.iterdir() if path.suffix.lower() in {".xlsx", ".csv"}])
    total_rows = 0

    with open(OUTPUT_PATH, "w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=OUTPUT_COLUMNS)
        writer.writeheader()

        for path in files:
            row_count = 0
            row_iter = stream_csv_rows(path) if path.suffix.lower() == ".csv" else stream_xlsx_rows(path)

            for raw_row in row_iter:
                normalized = normalize_row(raw_row, path)
                if normalized["address"] in (None, ""):
                    continue
                writer.writerow(normalized)
                row_count += 1

            total_rows += row_count
            print(f"{path.name}: {row_count} rows")

    print(f"\nSaved: {OUTPUT_PATH}")
    print(f"Rows: {total_rows}")


if __name__ == "__main__":
    main()
